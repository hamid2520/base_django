import os
import sys
from os import listdir
from os.path import isfile, isdir, join

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

directory = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

sys.path.insert(0, directory)
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'app.settings.local')

import django

django.setup()

from core.models import Exercise, Parameter, Option, Result, ExerciseData

def is_data_file(path):
    return isfile(path) and not path.split('/')[-1].startswith('~') and path.split('/')[-1].endswith('.xlsx')


def parse(file):
    workbook = load_workbook(filename=file)
    filename = file.split('/')[-1].replace('.xlsx', '')
    try:
        exercies = Exercise.objects.get(name=filename)
        print('EXERCISE INIT: ', exercies)
    except:
        print('لطفا تمرین را از قبل بسازید')
        return False

    sheet = workbook.active

    headers = [str(cell.value) for cell in sheet[1]]
    p_ids = []
    p_obj = []
    example_index = headers.index('example')
    for index, header in enumerate(headers[:example_index]):

        print(sheet[2][index].value)
        parameter = Parameter.objects.create(
            name=header,
            exercise=exercies,
            error_threshold=float(sheet[2][index].value),
        )
        if header in headers[example_index:]:
            parameter.type = 1
        if sheet[1][index].font.b:
            parameter.is_independent = True
        parameter.save()
        # p_ids.append(parameter.id)
        p_ids.append(index)
        p_obj.append(parameter)
        print('PARAMETER INIT: ', header)
    for col, header in enumerate(headers[example_index + 1:]):
        index = headers[:example_index].index(header)
        options = [str(cell.value) for cell in sheet[get_column_letter(example_index + col + 2)]]
        for option in options[1:]:
            if option == 'None':
                continue
            opt = Option.objects.create(
                name=option,
                parameter=p_obj[index],
            )
            # opt.parameter_id = p_ids[index]
            print('OPTION INIT: ', option)
    for i in range(2, sheet.max_row):
        result = Result.objects.create(
            title=i,
            is_example=str(sheet[i + 1][example_index].value) == '1',
            exercise=exercies,
        )
        print('RESULT INIT: ', i, str(sheet[i + 1][example_index].value) == '1')
        for index, obj in enumerate(p_obj):
            answer = ExerciseData.objects.create(
                value=str(sheet[i + 1][index].value),
                parameter=obj,
                result=result,
            )
            print('ANSWER INIT: ', str(sheet[i + 1][index].value))


if len(sys.argv) < 2:
    print('type the source file directory path, like: python3 exercise-data-entry.py /tmp')
    exit()
else:
    path = sys.argv[1]

if isdir(path):
    files = [join(path, f) for f in listdir(path) if is_data_file(join(path, f))]
    for f in files:
        # parse(f.split('/')[-1])
        parse(f)
